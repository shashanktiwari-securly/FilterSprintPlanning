"""
Build a monthly sprint-comparison workbook for the FILTER scrum team.

Discovers every ``product_FILTER/<sprint>-sprint-data/`` folder that has an
``issues.json`` file, computes the same metrics ``build_sprint_analysis.py``
uses (date-aware roster, India 2026 holidays, FILTER+PTGM+FDSE scope), groups
them by **anchor month** (the calendar month containing the most weekdays of
the sprint window — ties go to the earlier month), and writes:

    FILTER Monthly Sprint Comparison.xlsx

Tabs:

    1. Sprint Comparison   — one row per sprint with capacity, delivery,
                             completion, mid-sprint, P1 escalation. Monthly
                             subtotal rows separate the months.
    2. Monthly Rollup      — capacity + delivery aggregates per month and
                             month-over-month deltas.
    3. Priority Trends     — completed parents and hours per priority per
                             sprint, ordered chronologically.
    4. P1 & Mid-Sprint     — sprint-level snapshot of escalations and adds.

Usage::

    python build_monthly_comparison.py
    python build_monthly_comparison.py --out my_compare.xlsx
"""
from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import sprint_metrics as sm

ROOT = Path(__file__).resolve().parent

JIRA_BASE = "https://securly.atlassian.net/browse/"

GREEN = PatternFill("solid", fgColor="FFC6EFCE")
AMBER = PatternFill("solid", fgColor="FFFFEB9C")
RED = PatternFill("solid", fgColor="FFFFC7CE")
HEADER_FILL = PatternFill("solid", fgColor="FFD9E1F2")
HEADER_BAND_FILL = PatternFill("solid", fgColor="FF1F4E78")
HEADER_BAND_FONT = Font(bold=True, color="FFFFFFFF")
HEADER_FONT = Font(bold=True)
SUBTOTAL_FILL = PatternFill("solid", fgColor="FFEAEFF7")
SUBTOTAL_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14)
LINK_FONT = Font(color="0563C1", underline="single")


def rating_fill(value: float) -> PatternFill:
    if value >= 0.85:
        return GREEN
    if value >= 0.70:
        return AMBER
    return RED


def rating_label(value: float) -> str:
    if value >= 0.85:
        return "On Target"
    if value >= 0.70:
        return "Below Target"
    return "Significantly Below"


def safe_avg(values: list[float]) -> float:
    return sum(values) / len(values) if values else 0.0


def write_header(ws, headers: list[str], row: int = 1) -> None:
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[row].height = 36


def auto_width(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_sprint_comparison(ws, rows: list[dict]) -> None:
    title = (
        "Sprint Comparison — FILTER (FILTER + PTGM + FDSE) — "
        f"{len(rows)} sprints across {len({r['anchor_month_label'] for r in rows})} months"
    )
    ws.cell(1, 1, title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=18)

    headers = [
        "Anchor Month",
        "Sprint",
        "Window",
        "Working Days",
        "Holidays",
        "Effective WD",
        "Active Team",
        "Available (h)",
        "Net Available (h)",
        "Completed (h)",
        "Completed Tickets",
        "Spillover (OE h)",
        "Spillover Tickets",
        "Removed (OE h)",
        "Sprint Completion (tickets)",
        "Sprint Completion (planned hrs)",
        "Mid-sprint Adds (core)",
        "P1 Escalations (open / total)",
    ]
    write_header(ws, headers, row=3)

    rows_sorted = sorted(rows, key=lambda r: r["start"])
    by_month: dict[tuple[int, int], list[dict]] = {}
    for r in rows_sorted:
        by_month.setdefault((r["anchor_year"], r["anchor_month"]), []).append(r)

    r_idx = 4
    for ym in sorted(by_month):
        month_rows = by_month[ym]
        for r in month_rows:
            window = f"{r['start']:%d %b} → {r['end']:%d %b %Y}"
            ws.cell(r_idx, 1, r["anchor_month_label"])
            ws.cell(r_idx, 2, r["sprint_name"])
            ws.cell(r_idx, 3, window)
            ws.cell(r_idx, 4, r["working_days"]).number_format = "0"
            ws.cell(r_idx, 5, r["company_holidays"]).number_format = "0"
            ws.cell(r_idx, 6, r["effective_working_days"]).number_format = "0"
            ws.cell(r_idx, 7, r["active_team"]).number_format = "0"
            ws.cell(r_idx, 8, r["available_hrs"]).number_format = "0.00"
            ws.cell(r_idx, 9, r["net_available_hrs"]).number_format = "0.00"
            ws.cell(r_idx, 10, r["completed_hrs"]).number_format = "0.00"
            ws.cell(r_idx, 11, r["completed_records"]).number_format = "0"
            ws.cell(r_idx, 12, r["spillover_oe_hrs"]).number_format = "0.00"
            ws.cell(r_idx, 13, r["spillover_records"]).number_format = "0"
            ws.cell(r_idx, 14, r["removed_oe_hrs"]).number_format = "0.00"

            ct = ws.cell(r_idx, 15, r["completion_tickets"])
            ct.number_format = "0.0%"
            ct.fill = rating_fill(r["completion_tickets"])

            ch = ws.cell(r_idx, 16, r["completion_hours"])
            ch.number_format = "0.0%"
            ch.fill = rating_fill(r["completion_hours"])

            ws.cell(r_idx, 17, r["midsprint_additions_core"]).number_format = "0"
            ws.cell(
                r_idx, 18,
                f"{r['p1_escalations_open']} / {r['p1_escalations_total']}",
            )
            r_idx += 1

        # Subtotal row for the month
        ws.cell(r_idx, 1, f"{month_rows[0]['anchor_month_label']} TOTAL").font = SUBTOTAL_FONT
        ws.cell(r_idx, 2, f"{len(month_rows)} sprint(s)").font = SUBTOTAL_FONT
        ws.cell(r_idx, 3, "")
        ws.cell(r_idx, 4, sum(r["working_days"] for r in month_rows)).number_format = "0"
        ws.cell(r_idx, 5, sum(r["company_holidays"] for r in month_rows)).number_format = "0"
        ws.cell(r_idx, 6, sum(r["effective_working_days"] for r in month_rows)).number_format = "0"
        ws.cell(r_idx, 7, "").number_format = "0"
        ws.cell(r_idx, 8, sum(r["available_hrs"] for r in month_rows)).number_format = "0.00"
        ws.cell(r_idx, 9, sum(r["net_available_hrs"] for r in month_rows)).number_format = "0.00"
        ws.cell(r_idx, 10, sum(r["completed_hrs"] for r in month_rows)).number_format = "0.00"
        ws.cell(r_idx, 11, sum(r["completed_records"] for r in month_rows)).number_format = "0"
        ws.cell(r_idx, 12, sum(r["spillover_oe_hrs"] for r in month_rows)).number_format = "0.00"
        ws.cell(r_idx, 13, sum(r["spillover_records"] for r in month_rows)).number_format = "0"
        ws.cell(r_idx, 14, sum(r["removed_oe_hrs"] for r in month_rows)).number_format = "0.00"

        avg_ct = safe_avg([r["completion_tickets"] for r in month_rows])
        avg_ch = safe_avg([r["completion_hours"] for r in month_rows])
        ws.cell(r_idx, 15, avg_ct).number_format = "0.0%"
        ws.cell(r_idx, 15).fill = rating_fill(avg_ct)
        ws.cell(r_idx, 16, avg_ch).number_format = "0.0%"
        ws.cell(r_idx, 16).fill = rating_fill(avg_ch)
        ws.cell(r_idx, 17, sum(r["midsprint_additions_core"] for r in month_rows)).number_format = "0"
        ws.cell(
            r_idx, 18,
            f"{sum(r['p1_escalations_open'] for r in month_rows)} / "
            f"{sum(r['p1_escalations_total'] for r in month_rows)}",
        )
        for c in range(1, 19):
            if c not in (15, 16):
                ws.cell(r_idx, c).fill = SUBTOTAL_FILL
            ws.cell(r_idx, c).font = SUBTOTAL_FONT
        r_idx += 1

    auto_width(
        ws,
        [12, 18, 22, 9, 9, 9, 8, 11, 12, 11, 12, 11, 12, 11, 12, 13, 11, 16],
    )
    ws.freeze_panes = "C4"

    # ── Actual Delivery per Sprint (capacity vs delivery vs engagement) ─────
    block_row = r_idx + 2
    band = ws.cell(block_row, 1, "Actual Delivery per Sprint")
    band.font = HEADER_BAND_FONT
    band.fill = HEADER_BAND_FILL
    band.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.merge_cells(
        start_row=block_row, start_column=1, end_row=block_row, end_column=9
    )
    ws.row_dimensions[block_row].height = 22
    block_row += 1
    ws.cell(
        block_row, 1,
        "Ideal Capacity = effective working days × 8 × focus factor × active FTE. "
        "Delivered = sum of in-sprint Time Spent on the Completed bucket. "
        "Engaged = Delivered + Time Spent on the Spillover bucket (work in progress). "
        "Gap = Ideal − Delivered (positive ⇒ under-delivered, negative ⇒ over-delivered).",
    ).font = Font(italic=True, color="FF555555")
    ws.merge_cells(
        start_row=block_row, start_column=1, end_row=block_row, end_column=9
    )
    ws.row_dimensions[block_row].height = 30
    ws.cell(block_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    block_row += 1

    actual_headers = [
        "Sprint",
        "Ideal Capacity (hrs)",
        "Delivered — Completed (hrs)",
        "Engaged — Completed + NC (hrs)",
        "Delivery vs Ideal",
        "Engagement vs Ideal",
        "Gap (Ideal − Delivered)",
        "Rating",
        "Notes",
    ]
    write_header(ws, actual_headers, row=block_row)
    block_row += 1

    rows_chrono = sorted(rows, key=lambda r: r["start"])
    by_month_chrono: dict[tuple[int, int], list[dict]] = {}
    for r in rows_chrono:
        by_month_chrono.setdefault((r["anchor_year"], r["anchor_month"]), []).append(r)

    overall_ideal = 0.0
    overall_delivered = 0.0
    overall_engaged = 0.0

    for ym in sorted(by_month_chrono):
        month_rows = by_month_chrono[ym]
        for r in month_rows:
            ideal = r["ideal_capacity_hrs"]
            delivered = r["completed_hrs"]
            engaged = r["engaged_hrs"]
            gap = round(ideal - delivered, 2)
            dvi = (delivered / ideal) if ideal else 0
            evi = (engaged / ideal) if ideal else 0
            overall_ideal += ideal
            overall_delivered += delivered
            overall_engaged += engaged

            ws.cell(block_row, 1, r["sprint_name"]).font = HEADER_FONT
            ws.cell(block_row, 2, ideal).number_format = "0.00"
            ws.cell(block_row, 3, delivered).number_format = "0.00"
            ws.cell(block_row, 4, engaged).number_format = "0.00"
            dvi_cell = ws.cell(block_row, 5, dvi)
            dvi_cell.number_format = "0.0%"
            dvi_cell.fill = rating_fill(dvi)
            dvi_cell.font = HEADER_FONT
            evi_cell = ws.cell(block_row, 6, evi)
            evi_cell.number_format = "0.0%"
            ws.cell(block_row, 7, gap).number_format = "+0.00;-0.00;0.00"
            rating = rating_label(dvi)
            r_cell = ws.cell(block_row, 8, rating)
            r_cell.font = HEADER_FONT
            r_cell.fill = rating_fill(dvi)
            note_pieces: list[str] = []
            if r["inactive_excluded"]:
                note_pieces.append(
                    "Inactive: " + ", ".join(r["inactive_excluded"])
                )
            if r["company_holidays"]:
                note_pieces.append(
                    f"{r['company_holidays']} company holiday(s) in window"
                )
            if r["midsprint_additions_core"]:
                note_pieces.append(
                    f"{r['midsprint_additions_core']} mid-sprint adds"
                )
            ws.cell(block_row, 9, " · ".join(note_pieces))
            block_row += 1

        # Month subtotal row inside the Actual Delivery block
        ideal_sum = sum(r["ideal_capacity_hrs"] for r in month_rows)
        delivered_sum = sum(r["completed_hrs"] for r in month_rows)
        engaged_sum = sum(r["engaged_hrs"] for r in month_rows)
        gap_sum = round(ideal_sum - delivered_sum, 2)
        dvi_avg = (delivered_sum / ideal_sum) if ideal_sum else 0
        evi_avg = (engaged_sum / ideal_sum) if ideal_sum else 0
        ws.cell(
            block_row, 1, f"{month_rows[0]['anchor_month_label']} subtotal"
        ).font = SUBTOTAL_FONT
        ws.cell(block_row, 2, ideal_sum).number_format = "0.00"
        ws.cell(block_row, 3, delivered_sum).number_format = "0.00"
        ws.cell(block_row, 4, engaged_sum).number_format = "0.00"
        ws.cell(block_row, 5, dvi_avg).number_format = "0.0%"
        ws.cell(block_row, 5).fill = rating_fill(dvi_avg)
        ws.cell(block_row, 6, evi_avg).number_format = "0.0%"
        ws.cell(block_row, 7, gap_sum).number_format = "+0.00;-0.00;0.00"
        ws.cell(block_row, 8, rating_label(dvi_avg))
        ws.cell(block_row, 8).fill = rating_fill(dvi_avg)
        ws.cell(block_row, 9, f"{len(month_rows)} sprint(s)")
        for c in range(1, 10):
            if c not in (5, 8):
                ws.cell(block_row, c).fill = SUBTOTAL_FILL
            ws.cell(block_row, c).font = SUBTOTAL_FONT
        block_row += 1

    # Grand-total row
    gap_total = round(overall_ideal - overall_delivered, 2)
    dvi_total = (overall_delivered / overall_ideal) if overall_ideal else 0
    evi_total = (overall_engaged / overall_ideal) if overall_ideal else 0
    ws.cell(block_row, 1, "Total").font = HEADER_FONT
    ws.cell(block_row, 2, overall_ideal).number_format = "0.00"
    ws.cell(block_row, 3, overall_delivered).number_format = "0.00"
    ws.cell(block_row, 4, overall_engaged).number_format = "0.00"
    ws.cell(block_row, 5, dvi_total).number_format = "0.0%"
    ws.cell(block_row, 5).fill = rating_fill(dvi_total)
    ws.cell(block_row, 6, evi_total).number_format = "0.0%"
    ws.cell(block_row, 7, gap_total).number_format = "+0.00;-0.00;0.00"
    ws.cell(block_row, 8, rating_label(dvi_total))
    ws.cell(block_row, 8).fill = rating_fill(dvi_total)
    ws.cell(block_row, 9, f"{len(rows)} sprint(s) across {len(by_month_chrono)} month(s)")
    for c in range(1, 10):
        ws.cell(block_row, c).font = HEADER_FONT
        if c not in (5, 8):
            ws.cell(block_row, c).fill = HEADER_FILL


def build_monthly_rollup(ws, rows: list[dict]) -> None:
    ws.cell(1, 1, "Monthly Rollup — capacity & delivery").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws.cell(
        2, 1,
        "Aggregated across the sprints anchored in each calendar month. "
        "Completion % values are averaged across the sprints in the month "
        "(equal-weighted, not capacity-weighted).",
    ).font = Font(italic=True, color="FF555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
    ws.row_dimensions[2].height = 28
    ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="top")

    headers = [
        "Month",
        "Sprints",
        "Working Days",
        "Holidays",
        "Available (h)",
        "Net Available (h)",
        "Completed (h)",
        "Spillover (OE h)",
        "Removed (OE h)",
        "Avg Completion (tickets)",
        "Avg Completion (planned hrs)",
        "Mid-sprint Adds",
    ]
    write_header(ws, headers, row=4)

    rows_sorted = sorted(rows, key=lambda r: (r["anchor_year"], r["anchor_month"], r["start"]))
    by_month: dict[tuple[int, int], list[dict]] = {}
    for r in rows_sorted:
        by_month.setdefault((r["anchor_year"], r["anchor_month"]), []).append(r)

    r_idx = 5
    monthly_summaries: list[dict] = []
    for ym in sorted(by_month):
        month_rows = by_month[ym]
        label = month_rows[0]["anchor_month_label"]
        avg_ct = safe_avg([r["completion_tickets"] for r in month_rows])
        avg_ch = safe_avg([r["completion_hours"] for r in month_rows])
        completed_hrs = round(sum(r["completed_hrs"] for r in month_rows), 2)
        net_available = round(sum(r["net_available_hrs"] for r in month_rows), 2)
        summary = {
            "label": label,
            "sprints": len(month_rows),
            "working_days": sum(r["working_days"] for r in month_rows),
            "holidays": sum(r["company_holidays"] for r in month_rows),
            "available_hrs": round(sum(r["available_hrs"] for r in month_rows), 2),
            "net_available_hrs": net_available,
            "completed_hrs": completed_hrs,
            "spillover_oe": round(sum(r["spillover_oe_hrs"] for r in month_rows), 2),
            "removed_oe": round(sum(r["removed_oe_hrs"] for r in month_rows), 2),
            "avg_completion_tickets": avg_ct,
            "avg_completion_hours": avg_ch,
            "midsprint": sum(r["midsprint_additions_core"] for r in month_rows),
        }
        monthly_summaries.append(summary)

        ws.cell(r_idx, 1, label).font = SUBTOTAL_FONT
        ws.cell(r_idx, 2, summary["sprints"]).number_format = "0"
        ws.cell(r_idx, 3, summary["working_days"]).number_format = "0"
        ws.cell(r_idx, 4, summary["holidays"]).number_format = "0"
        ws.cell(r_idx, 5, summary["available_hrs"]).number_format = "0.00"
        ws.cell(r_idx, 6, summary["net_available_hrs"]).number_format = "0.00"
        ws.cell(r_idx, 7, summary["completed_hrs"]).number_format = "0.00"
        ws.cell(r_idx, 8, summary["spillover_oe"]).number_format = "0.00"
        ws.cell(r_idx, 9, summary["removed_oe"]).number_format = "0.00"
        ct = ws.cell(r_idx, 10, avg_ct); ct.number_format = "0.0%"; ct.fill = rating_fill(avg_ct)
        ch = ws.cell(r_idx, 11, avg_ch); ch.number_format = "0.0%"; ch.fill = rating_fill(avg_ch)
        ws.cell(r_idx, 12, summary["midsprint"]).number_format = "0"
        r_idx += 1

    # Month-over-month delta rows
    if len(monthly_summaries) >= 2:
        r_idx += 1
        ws.cell(r_idx, 1, "Month-over-month delta").font = HEADER_FONT
        ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=12)
        r_idx += 1
        delta_headers = [
            "From → To",
            "Δ Sprints",
            "Δ Working Days",
            "Δ Holidays",
            "Δ Available (h)",
            "Δ Net Avail (h)",
            "Δ Completed (h)",
            "Δ Spillover (h)",
            "Δ Removed (h)",
            "Δ Avg Compl. (tickets)",
            "Δ Avg Compl. (hrs)",
            "Δ Mid-sprint Adds",
        ]
        write_header(ws, delta_headers, row=r_idx)
        r_idx += 1
        for prev, cur in zip(monthly_summaries, monthly_summaries[1:]):
            ws.cell(r_idx, 1, f"{prev['label']} → {cur['label']}").font = HEADER_FONT
            ws.cell(r_idx, 2, cur["sprints"] - prev["sprints"]).number_format = "+0;-0;0"
            ws.cell(r_idx, 3, cur["working_days"] - prev["working_days"]).number_format = "+0;-0;0"
            ws.cell(r_idx, 4, cur["holidays"] - prev["holidays"]).number_format = "+0;-0;0"
            ws.cell(
                r_idx, 5, round(cur["available_hrs"] - prev["available_hrs"], 2)
            ).number_format = "+0.00;-0.00;0.00"
            ws.cell(
                r_idx, 6, round(cur["net_available_hrs"] - prev["net_available_hrs"], 2)
            ).number_format = "+0.00;-0.00;0.00"
            ws.cell(
                r_idx, 7, round(cur["completed_hrs"] - prev["completed_hrs"], 2)
            ).number_format = "+0.00;-0.00;0.00"
            ws.cell(
                r_idx, 8, round(cur["spillover_oe"] - prev["spillover_oe"], 2)
            ).number_format = "+0.00;-0.00;0.00"
            ws.cell(
                r_idx, 9, round(cur["removed_oe"] - prev["removed_oe"], 2)
            ).number_format = "+0.00;-0.00;0.00"
            ws.cell(
                r_idx, 10,
                round(cur["avg_completion_tickets"] - prev["avg_completion_tickets"], 4),
            ).number_format = "+0.0%;-0.0%;0.0%"
            ws.cell(
                r_idx, 11,
                round(cur["avg_completion_hours"] - prev["avg_completion_hours"], 4),
            ).number_format = "+0.0%;-0.0%;0.0%"
            ws.cell(r_idx, 12, cur["midsprint"] - prev["midsprint"]).number_format = "+0;-0;0"
            r_idx += 1

    auto_width(ws, [16, 9, 11, 9, 11, 11, 11, 12, 11, 16, 16, 14])
    ws.freeze_panes = "B5"


def build_priority_trends(ws, rows: list[dict]) -> None:
    ws.cell(1, 1, "Priority Trends — Completed by Priority across sprints").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws.cell(
        2, 1,
        "Tickets (parents) — subtasks rolled up to their parent. "
        "Hours include parents + every subtask (matches per-person Completed total). "
        "Ticket key lists at the bottom show the actual parent keys for reference.",
    ).font = Font(italic=True, color="FF555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
    ws.row_dimensions[2].height = 32
    ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="top")

    rows_sorted = sorted(rows, key=lambda r: r["start"])
    priorities = ["P0", "P1", "P2", "P3", "P4"]
    headers = ["Anchor Month", "Sprint", "Window", "Bucket"] + [f"{p} (parents)" for p in priorities] + ["Total Parents"]
    write_header(ws, headers, row=4)
    r_idx = 5
    for r in rows_sorted:
        window = f"{r['start']:%d %b} → {r['end']:%d %b}"
        ws.cell(r_idx, 1, r["anchor_month_label"])
        ws.cell(r_idx, 2, r["sprint_name"])
        ws.cell(r_idx, 3, window)
        ws.cell(r_idx, 4, "Tickets (parents)")
        total = 0
        for c, p in enumerate(priorities, start=5):
            v = r["priority"].get(p, {}).get("count", 0)
            ws.cell(r_idx, c, v).number_format = "0"
            total += v
        ws.cell(r_idx, 4 + len(priorities) + 1, total).number_format = "0"
        ws.cell(r_idx, 4 + len(priorities) + 1).font = HEADER_FONT
        r_idx += 1

    r_idx += 1
    ws.cell(r_idx, 1, "Hours (incl. subtasks)").font = HEADER_FONT
    ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=12)
    r_idx += 1
    headers_h = ["Anchor Month", "Sprint", "Window", "Bucket"] + [f"{p} (hrs)" for p in priorities] + ["Total Hrs"]
    write_header(ws, headers_h, row=r_idx)
    r_idx += 1
    for r in rows_sorted:
        window = f"{r['start']:%d %b} → {r['end']:%d %b}"
        ws.cell(r_idx, 1, r["anchor_month_label"])
        ws.cell(r_idx, 2, r["sprint_name"])
        ws.cell(r_idx, 3, window)
        ws.cell(r_idx, 4, "Hours (incl. subtasks)")
        total = 0.0
        for c, p in enumerate(priorities, start=5):
            v = r["priority"].get(p, {}).get("hours", 0.0)
            ws.cell(r_idx, c, v).number_format = "0.00"
            total += v
        ws.cell(r_idx, 4 + len(priorities) + 1, round(total, 2)).number_format = "0.00"
        ws.cell(r_idx, 4 + len(priorities) + 1).font = HEADER_FONT
        r_idx += 1

    # Reference — ticket keys per priority per sprint
    r_idx += 2
    band = ws.cell(r_idx, 1, "Reference — Completed parent ticket keys (Jira)")
    band.font = HEADER_BAND_FONT
    band.fill = HEADER_BAND_FILL
    band.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=12)
    ws.row_dimensions[r_idx].height = 22
    r_idx += 1
    ref_headers = ["Anchor Month", "Sprint", "Priority", "Parent Tickets (semicolon-separated)"]
    write_header(ws, ref_headers, row=r_idx)
    r_idx += 1
    for r in rows_sorted:
        for p in priorities:
            keys = r["priority"].get(p, {}).get("keys", []) or []
            if not keys:
                continue
            ws.cell(r_idx, 1, r["anchor_month_label"])
            ws.cell(r_idx, 2, r["sprint_name"])
            ws.cell(r_idx, 3, p)
            cell = ws.cell(r_idx, 4, "; ".join(keys))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(
                start_row=r_idx, start_column=4, end_row=r_idx, end_column=12
            )
            row_height = max(15, 15 * (1 + len(keys) // 8))
            ws.row_dimensions[r_idx].height = min(row_height, 90)
            r_idx += 1

    auto_width(ws, [12, 18, 22, 18, 11, 11, 11, 11, 11, 13])
    ws.freeze_panes = "C5"


def build_p1_and_midsprint(ws, rows: list[dict]) -> None:
    ws.cell(1, 1, "P1 Escalations & Mid-sprint additions — sprint snapshot").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    headers = [
        "Anchor Month",
        "Sprint",
        "Window",
        "P1 Escalations (total)",
        "P1 Escalations (open)",
        "Mid-sprint Adds (core)",
        "Spillover Tickets",
        "Removed Tickets",
    ]
    write_header(ws, headers, row=3)
    rows_sorted = sorted(rows, key=lambda r: r["start"])
    r_idx = 4
    for r in rows_sorted:
        window = f"{r['start']:%d %b} → {r['end']:%d %b}"
        ws.cell(r_idx, 1, r["anchor_month_label"])
        ws.cell(r_idx, 2, r["sprint_name"])
        ws.cell(r_idx, 3, window)
        ws.cell(r_idx, 4, r["p1_escalations_total"]).number_format = "0"
        cell = ws.cell(r_idx, 5, r["p1_escalations_open"])
        cell.number_format = "0"
        if r["p1_escalations_open"] > 0:
            cell.fill = RED
            cell.font = Font(bold=True, color="FF9C0006")
        ws.cell(r_idx, 6, r["midsprint_additions_core"]).number_format = "0"
        ws.cell(r_idx, 7, r["spillover_records"]).number_format = "0"
        ws.cell(r_idx, 8, r["removed_records"]).number_format = "0"
        r_idx += 1

    # Reference — ticket keys for the metrics above
    r_idx += 2
    band = ws.cell(r_idx, 1, "Reference — Jira ticket keys per metric")
    band.font = HEADER_BAND_FONT
    band.fill = HEADER_BAND_FILL
    band.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=8)
    ws.row_dimensions[r_idx].height = 22
    r_idx += 1
    ref_headers = ["Anchor Month", "Sprint", "Category", "Tickets (semicolon-separated)"]
    write_header(ws, ref_headers, row=r_idx)
    r_idx += 1
    categories = [
        ("P1 Escalation — open", "p1_escalation_open_keys"),
        ("P1 Escalation — total", "p1_escalation_keys"),
        ("Mid-sprint additions (core)", "midsprint_added_keys"),
        ("Spillover (Not Completed)", "spillover_keys"),
        ("Removed", "removed_keys"),
    ]
    for r in rows_sorted:
        for label, attr in categories:
            keys = r.get(attr) or []
            if not keys:
                continue
            ws.cell(r_idx, 1, r["anchor_month_label"])
            ws.cell(r_idx, 2, r["sprint_name"])
            ws.cell(r_idx, 3, label)
            cell = ws.cell(r_idx, 4, "; ".join(keys))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(
                start_row=r_idx, start_column=4, end_row=r_idx, end_column=8
            )
            row_height = max(15, 15 * (1 + len(keys) // 8))
            ws.row_dimensions[r_idx].height = min(row_height, 120)
            if label.startswith("P1 Escalation — open") and keys:
                for c in range(1, 9):
                    ws.cell(r_idx, c).fill = RED
                ws.cell(r_idx, 3).font = Font(bold=True, color="FF9C0006")
            r_idx += 1

    auto_width(ws, [12, 18, 22, 60, 14, 14, 14, 14])
    ws.freeze_panes = "C4"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--out",
        default=None,
        help="Output workbook path (default: 'FILTER Monthly Sprint Comparison.xlsx')",
    )
    args = ap.parse_args()

    folders = sm.discover_sprints()
    if not folders:
        raise SystemExit(
            "No sprint data found under product_FILTER/. "
            "Run `python run_sprint_analyses.py` first."
        )
    rows: list[dict] = []
    for folder in folders:
        try:
            rows.append(sm.compute_metrics(folder))
        except Exception as exc:  # noqa: BLE001
            print(f"  warning: failed to compute metrics for {folder.name}: {exc}")
    if not rows:
        raise SystemExit("No metrics could be computed.")

    wb = Workbook()
    sc = wb.active
    sc.title = "Sprint Comparison"
    build_sprint_comparison(sc, rows)

    mr = wb.create_sheet("Monthly Rollup")
    build_monthly_rollup(mr, rows)

    pt = wb.create_sheet("Priority Trends")
    build_priority_trends(pt, rows)

    pm = wb.create_sheet("P1 & Mid-Sprint")
    build_p1_and_midsprint(pm, rows)

    out_path = ROOT / "FILTER Monthly Sprint Comparison.xlsx"
    if args.out:
        out_path = Path(args.out)
        if not out_path.is_absolute():
            out_path = ROOT / out_path
    try:
        wb.save(out_path)
    except PermissionError:
        suffix = datetime.now().strftime("%H%M%S")
        out_path = out_path.with_name(f"{out_path.stem}.{suffix}{out_path.suffix}")
        print(f"Primary file is open in Excel — saving to {out_path.name} instead.")
        wb.save(out_path)
    print(f"Wrote {out_path}")
    print(f"  Sprints analyzed: {len(rows)}")
    by_month: dict[str, int] = {}
    for r in rows:
        by_month[r["anchor_month_label"]] = by_month.get(r["anchor_month_label"], 0) + 1
    for label, count in by_month.items():
        print(f"    {label}: {count} sprint(s)")


if __name__ == "__main__":
    main()
