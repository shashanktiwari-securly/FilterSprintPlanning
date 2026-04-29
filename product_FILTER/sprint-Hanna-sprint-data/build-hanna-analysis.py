"""
Build FILTER Sprint Hanna analysis workbook (Cases-style layout).

Tabs (in order):
    1. Capacity vs Delivery        — capacity, delivery, completed-by-priority
    2. Per-person Delivery
    3. Completed
    4. Not Completed
    5. Removed
    6. P1 Escalation               — highlighted view of every Hanna ticket whose
                                     summary contains "P1 Escalation"

Inputs (regenerate with the listed scripts):

* ``sprint-meta.json``     ← ``python fetch_filter_sprint_meta.py sprint-Hanna``
* ``issues.json``          ← ``python fetch_filter_sprint_issues.py sprint-Hanna``
* per-person availability  ← ``compute_hanna_capacity.py`` (PTO ∩ sprint window)

Output: ``FILTER Sprint Hanna Analysis.xlsx`` at the workspace root.
"""
from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

import compute_hanna_capacity as cap  # noqa: E402

ROOT = Path(__file__).resolve().parents[2]
JIRA_BASE = "https://securly.atlassian.net/browse/"
SPRINT_NAME = "sprint-Hanna"
ISSUES_FILE = SCRIPT_DIR / "issues.json"

CORE_TEAM_ORDER = [
    "Altamash Heroli",
    "Amit Shete",
    "Amol Mithari",
    "Arun Thakur",
    "Ashish Modak",
    "Ashish Temurnikar",
    "Ayesha Kamde",
    "Harsh Verma",
    "Harshada Kude",
    "Prateek Fotedar",
    "Pratik Tiwari",
    "Rachit Mishra",
    "Ravi Raj",
    "Sagar Satpute",
    "Shailendra Singh",
    "Surabhi Choudhary",
    "Swaroop Chavhan",
    "Vaibhav Kumbhar",
]

# Per Securly-PjM-Skills.md Part 2 — concise role / function for the team-composition
# block used on the Capacity vs Delivery sheet.
TEAM_COMPOSITION = [
    ("Shashank Tiwari", "Project Manager", "PM / Delivery", 0.0, "Lead", "FTE",
     "Not counted in delivery capacity"),
    ("Harsh Verma", "FE", "FrontEnd Architect (Angular)", 1.0, "Senior", "FTE", ""),
    ("Arun Thakur", "FE + Ext", "Angular / Extension", 1.0, "Medium", "FTE", ""),
    ("Harshada Kude", "FE", "Angular", 1.0, "Junior", "FTE", ""),
    ("Shailendra Singh", "FE", "Angular", 1.0, "Medium", "FTE", ""),
    ("Swaroop Chavhan", "FE + Ext", "Angular / Extension", 1.0, "Medium", "FTE", ""),
    ("Ashish Modak", "BE Architect", "Extension / Go / PHP", 1.0, "Senior", "FTE", ""),
    ("Prateek Fotedar", "BE (www)", "Go / PHP", 1.0, "Senior", "FTE", ""),
    ("Sagar Satpute", "BE (www)", "Go / PHP", 1.0, "Medium", "FTE", ""),
    ("Vaibhav Kumbhar", "BE", "Go / PHP", 1.0, "Medium", "FTE",
     "Full sprint PTO (27 Apr–15 May)"),
    ("Ashish Temurnikar", "BE", "DNS / SPAC / Squid / Rust", 1.0, "Senior", "FTE", ""),
    ("Ravi Raj", "BE", "DNS / SPAC / Squid / Rust", 1.0, "Medium", "FTE", ""),
    ("Pratik Tiwari", "BE + Ext", "Extension / Go / Rust", 1.0, "Senior", "FTE", ""),
    ("Altamash Heroli", "QA", "Pri. Manual QA", 1.0, "Senior", "FTE", ""),
    ("Surabhi Choudhary", "QA", "Sr. Manual QA", 1.0, "Senior", "FTE", ""),
    ("Ayesha Kamde", "QA", "Manual QA", 1.0, "Junior", "FTE", ""),
    ("Amol Mithari", "Auto QA Lead", "Team Lead — Auto QA / SDET", 1.0, "Lead", "FTE", ""),
    ("Amit Shete", "Auto QA", "Pri. SDET", 1.0, "Senior", "FTE", ""),
    ("Rachit Mishra", "Auto QA", "SDET", 1.0, "Junior", "FTE", ""),
]

DETAIL_HEADERS = [
    "Issue Type",
    "Issue key",
    "Summary",
    "Parent key",
    "Parent type",
    "Assignee",
    "Status",
    "Resolution",
    "Priority",
    "Created",
    "Resolved",
    "Updated",
    "Sprint",
    "Time Spent",
    "Original Est",
    "Project",
    "Labels",
    "Blockers",
]

GREEN = PatternFill("solid", fgColor="FFC6EFCE")
AMBER = PatternFill("solid", fgColor="FFFFEB9C")
RED = PatternFill("solid", fgColor="FFFFC7CE")
INPUT_FILL = PatternFill("solid", fgColor="FFFFEB9C")
HEADER_FILL = PatternFill("solid", fgColor="FFD9E1F2")
HEADER_FONT = Font(bold=True)

P1_ESC_HEADER_FILL = PatternFill("solid", fgColor="FFC00000")
P1_ESC_HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=12)
P1_ESC_TITLE_FONT = Font(bold=True, size=14, color="FFC00000")
P1_ESC_OPEN_FILL = PatternFill("solid", fgColor="FFFFE5E5")

ADDED_MID_SPRINT_FILL = PatternFill("solid", fgColor="FFFFF2CC")
ADDED_MID_SPRINT_FONT = Font(bold=True, color="FF7F6000")

PRIORITY_ORDER = ["P0", "P1", "P2", "P3", "P4", "(none)"]
P1_ESC_PATTERN = re.compile(r"p1\s*escal", re.IGNORECASE)

MISC_BUFFER_PCT = 0.20  # blanket reservation for meetings, scrums, code review, etc.


def load_issues() -> dict:
    if not ISSUES_FILE.is_file():
        raise SystemExit(
            f"Missing {ISSUES_FILE}. Run from repo root:\n"
            "  python fetch_filter_sprint_issues.py sprint-Hanna"
        )
    return json.loads(ISSUES_FILE.read_text(encoding="utf-8"))


def in_core(name: str | None) -> bool:
    return bool(name) and name in CORE_TEAM_ORDER


def parse_date(s: str | None):
    if not s:
        return None
    return datetime.strptime(s, "%Y-%m-%d")


def detail_row(rec: dict) -> list:
    return [
        rec.get("issue_type") or "",
        rec.get("key") or "",
        rec.get("summary") or "",
        rec.get("parent_key") or "",
        rec.get("parent_type") or "",
        rec.get("assignee") or "",
        rec.get("status") or "",
        rec.get("resolution") or "",
        rec.get("priority") or "",
        parse_date(rec.get("created")),
        parse_date(rec.get("resolved")),
        parse_date(rec.get("updated")),
        rec.get("sprint") or "",
        float(rec.get("time_spent_hrs") or 0.0),
        float(rec.get("original_estimate_hrs") or 0.0),
        rec.get("project") or "",
        ", ".join(rec.get("labels") or []),
        rec.get("blockers") or "",
    ]


def write_detail_sheet(
    ws,
    records: list[dict],
    sort_keys: tuple[str, ...],
    flag_mid_sprint: bool = False,
):
    headers = list(DETAIL_HEADERS)
    if flag_mid_sprint:
        headers.append("Added Mid-Sprint")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    records = sorted(
        records,
        key=lambda r: tuple(
            (
                0 if r.get("added_during_sprint") and flag_mid_sprint else 1,
                *((r.get(k) or "") for k in sort_keys),
            )
        ),
    )
    n_total = len(records)
    n_mid_sprint = 0
    for r, rec in enumerate(records, 2):
        row = detail_row(rec)
        is_added = bool(rec.get("added_during_sprint"))
        if flag_mid_sprint and is_added:
            n_mid_sprint += 1
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            if c == 2 and rec.get("key"):
                cell.hyperlink = f"{JIRA_BASE}{rec['key']}"
                cell.font = Font(color="0563C1", underline="single")
            if c in (10, 11, 12) and isinstance(val, datetime):
                cell.number_format = "yyyy-mm-dd"
            if c in (14, 15):
                cell.number_format = "0.00"
            if c == 3:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if flag_mid_sprint and is_added:
                cell.fill = ADDED_MID_SPRINT_FILL
        if flag_mid_sprint:
            tag = ws.cell(r, len(DETAIL_HEADERS) + 1, "YES" if is_added else "")
            if is_added:
                tag.font = ADDED_MID_SPRINT_FONT
                tag.fill = ADDED_MID_SPRINT_FILL
    widths = [12, 16, 60, 16, 14, 22, 18, 14, 10, 12, 12, 12, 18, 12, 12, 10, 18, 14]
    if flag_mid_sprint:
        widths.append(16)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    if records:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{n_total + 1}"
    return n_mid_sprint


def per_person_totals(records: list[dict], field: str) -> dict[str, float]:
    out: dict[str, float] = {n: 0.0 for n in CORE_TEAM_ORDER}
    for rec in records:
        a = rec.get("assignee")
        if a in out:
            out[a] = round(out[a] + float(rec.get(field) or 0.0), 2)
    return out


def priority_key(rec: dict) -> str:
    p = rec.get("priority") or ""
    return p if p in PRIORITY_ORDER else "(none)"


SUBTASK_TYPES: set[str] = {
    "Dev",
    "QA",
    "Automation",
    "Review",
    "Test Case Creation",
    "Debug",
}


def is_subtask(rec: dict) -> bool:
    """Filter convention: Dev/QA/Automation/Review/Test Case Creation/Debug are
    subtasks under a parent Story/Task/Defect/Escape Defect. See Securly-PjM-Skills
    Part 2 — Jira hierarchy."""
    return (rec.get("issue_type") or "") in SUBTASK_TYPES


def priority_breakdown(
    records: list[dict],
    by_key: dict[str, dict] | None = None,
) -> dict[str, dict[str, float]]:
    """{priority: {count, hours}} for a set of completed records.

    Counts roll subtasks up to their parent so a Story + 3 subtasks = 1 ticket.
    Hours sum every record (parent + subtasks) so total effort is preserved and
    matches the per-person Completed total. Priority is taken from the parent
    (PM-owned in Filter convention) with the subtask priority as fallback."""
    by_key = by_key or {}
    out = {
        p: {"count": 0, "count_all": 0, "hours": 0.0, "_keys": set()}
        for p in PRIORITY_ORDER
    }
    out["(none)"] = {"count": 0, "count_all": 0, "hours": 0.0, "_keys": set()}
    for rec in records:
        rollup_key = (
            rec.get("parent_key")
            if is_subtask(rec) and rec.get("parent_key")
            else rec.get("key")
        )
        parent_rec = by_key.get(rollup_key) if rollup_key else None
        pri_source = parent_rec or rec
        pri = pri_source.get("priority") or rec.get("priority") or ""
        bucket = out.get(pri) or out["(none)"]
        if rollup_key and rollup_key not in bucket["_keys"]:
            bucket["_keys"].add(rollup_key)
            bucket["count"] += 1
        bucket["count_all"] += 1
        bucket["hours"] = round(bucket["hours"] + float(rec.get("time_spent_hrs") or 0.0), 2)
    for v in out.values():
        v.pop("_keys", None)
    return {k: v for k, v in out.items() if k in PRIORITY_ORDER}


def is_p1_escalation(rec: dict) -> bool:
    """Title contains 'P1 Escalation' AND issue type = Escape Defect AND priority = P1."""
    if not P1_ESC_PATTERN.search(rec.get("summary") or ""):
        return False
    if (rec.get("issue_type") or "") != "Escape Defect":
        return False
    if (rec.get("priority") or "") != "P1":
        return False
    return True


def write_p1_escalation_sheet(ws, records: list[dict], sprint_meta: dict) -> int:
    title = (
        f"P1 Escalation — FILTER · {sprint_meta['sprint_name']} "
        f"({sprint_meta['start_date']} → {sprint_meta['end_date']})"
    )
    ws.cell(1, 1, title).font = P1_ESC_TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(DETAIL_HEADERS))

    subtitle = (
        f"Hanna tickets where Issue Type = Escape Defect AND Priority = P1 AND "
        f"summary contains 'P1 Escalation' (regardless of bucket). Total = {len(records)}."
    )
    ws.cell(2, 1, subtitle).font = Font(italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(DETAIL_HEADERS))

    header_row = 4
    extra = ["Bucket"]
    headers_full = DETAIL_HEADERS + extra
    for c, h in enumerate(headers_full, 1):
        cell = ws.cell(header_row, c, h)
        cell.font = P1_ESC_HEADER_FONT
        cell.fill = P1_ESC_HEADER_FILL

    records_sorted = sorted(
        records,
        key=lambda r: (
            0 if (r.get("status_category") or "") != "done" else 1,
            r.get("priority") or "ZZ",
            r.get("assignee") or "",
            r.get("key") or "",
        ),
    )

    for r, rec in enumerate(records_sorted, header_row + 1):
        row = detail_row(rec)
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            if c == 2 and rec.get("key"):
                cell.hyperlink = f"{JIRA_BASE}{rec['key']}"
                cell.font = Font(color="0563C1", underline="single")
            if c in (10, 11, 12) and isinstance(val, datetime):
                cell.number_format = "yyyy-mm-dd"
            if c in (14, 15):
                cell.number_format = "0.00"
            if c == 3:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        bucket_label = rec.get("_bucket", "")
        bcell = ws.cell(r, len(DETAIL_HEADERS) + 1, bucket_label)
        if (rec.get("status_category") or "") != "done":
            for c in range(1, len(headers_full) + 1):
                ws.cell(r, c).fill = P1_ESC_OPEN_FILL
            bcell.font = Font(bold=True, color="FFC00000")

    widths = [12, 16, 60, 16, 14, 22, 18, 14, 10, 12, 12, 12, 18, 12, 12, 10, 18, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{header_row + 1}"
    if records_sorted:
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(len(headers_full))}{header_row + len(records_sorted)}"
        )
    return len(records_sorted)


def main() -> None:
    payload = load_issues()
    sprint_meta = payload["sprint_meta"]
    issues = payload["issues"]
    bucket = payload["bucket"]

    sprint_start = date.fromisoformat(sprint_meta["start_date"])
    sprint_end = date.fromisoformat(sprint_meta["end_date"])
    sprint_wd = int(sprint_meta["working_days"])
    avail_hours = cap.compute_availability_hours(sprint_start, sprint_end, sprint_wd)
    leave_days = cap.compute_leave_days(sprint_start, sprint_end)

    by_key = {r["key"]: r for r in issues}
    completed_recs = [by_key[k] for k in bucket["completed"] if k in by_key and in_core(by_key[k].get("assignee"))]
    not_completed_recs = [by_key[k] for k in bucket["not_completed"] if k in by_key and in_core(by_key[k].get("assignee"))]
    removed_recs = [by_key[k] for k in bucket["removed"] if k in by_key and in_core(by_key[k].get("assignee"))]

    completed_hrs_by = per_person_totals(completed_recs, "time_spent_hrs")
    notcomp_hrs_by = per_person_totals(not_completed_recs, "original_estimate_hrs")
    removed_hrs_by = per_person_totals(removed_recs, "original_estimate_hrs")

    completed_priority = priority_breakdown(completed_recs, by_key)

    def _sum_oe(records: list[dict]) -> float:
        return round(sum(float(r.get("original_estimate_hrs") or 0.0) for r in records), 2)

    def _sum_ts(records: list[dict]) -> float:
        return round(sum(float(r.get("time_spent_hrs") or 0.0) for r in records), 2)

    n_completed = len(completed_recs)
    n_spillover = len(not_completed_recs)
    n_removed = len(removed_recs)
    completed_oe_hrs = _sum_oe(completed_recs)
    completed_ts_hrs = _sum_ts(completed_recs)
    spillover_oe_hrs = _sum_oe(not_completed_recs)
    removed_oe_hrs = _sum_oe(removed_recs)

    planned_completed = [r for r in completed_recs if not r.get("added_during_sprint")]
    planned_spillover = [r for r in not_completed_recs if not r.get("added_during_sprint")]
    n_planned_completed = len(planned_completed)
    n_planned_spillover = len(planned_spillover)
    planned_completed_oe = _sum_oe(planned_completed)
    planned_spillover_oe = _sum_oe(planned_spillover)

    n_added_total_core = sum(
        1 for r in issues if r.get("added_during_sprint") and in_core(r.get("assignee"))
    )

    bucket_keys = {
        "Completed": set(bucket["completed"]),
        "Not Completed": set(bucket["not_completed"]),
        "Removed": set(bucket["removed"]),
    }
    p1_escalation_recs: list[dict] = []
    for rec in issues:
        if not is_p1_escalation(rec):
            continue
        k = rec["key"]
        b = next((name for name, keys in bucket_keys.items() if k in keys), "Other")
        rec_with_bucket = dict(rec)
        rec_with_bucket["_bucket"] = b
        p1_escalation_recs.append(rec_with_bucket)

    wb = Workbook()

    cap_ws = wb.active
    cap_ws.title = "Capacity vs Delivery"

    title = (
        f"Capacity vs Delivery — FILTER · {SPRINT_NAME} "
        f"({sprint_start.strftime('%a %d %b')} → {sprint_end.strftime('%a %d %b %Y')})"
    )
    cap_ws.cell(1, 1, title).font = Font(bold=True, size=14)
    cap_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    cap_ws.cell(2, 1, "Team Composition (per Securly-PjM-Skills.md Part 2 — Filter India roster)").font = Font(bold=True)
    cap_ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

    comp_headers = ["Name", "Role", "Function / Split", "Allocation (FTE)", "Seniority", "Employment", "Notes"]
    for c, h in enumerate(comp_headers, 1):
        cell = cap_ws.cell(3, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row_index = 4
    for member in TEAM_COMPOSITION:
        for c, val in enumerate(member, 1):
            cell = cap_ws.cell(row_index, c, val)
            if c == 4:
                cell.number_format = "0.0"
        row_index += 1
    total_alloc_row = row_index
    cap_ws.cell(total_alloc_row, 1, "Total FILTER allocation").font = Font(bold=True)
    cap_ws.cell(total_alloc_row, 4, f"=SUM(D4:D{total_alloc_row-1})").number_format = "0.0"
    cap_ws.cell(total_alloc_row, 5, "FTE-equivalent heads available to FILTER")

    row_index = total_alloc_row + 2
    cap_ws.cell(row_index, 1, "Capacity Parameters (yellow cells are inputs — re-run the script after edits)").font = Font(bold=True)
    cap_ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=7)
    row_index += 1
    for c, h in enumerate(["Parameter", "Value", "Description"], 1):
        cell = cap_ws.cell(row_index, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row_index += 1

    company_holidays = cap.company_holidays_in_sprint(sprint_start, sprint_end)
    holiday_count = len(company_holidays)
    holiday_descr = (
        ", ".join(f"{d:%d-%b}: {name}" for d, name in company_holidays) or "none in window"
    )
    eff_sprint_wd = max(0, sprint_wd - holiday_count)

    p_rows = {}
    p_rows["wd"] = row_index
    cap_ws.cell(row_index, 1, "Sprint length (working days)")
    c = cap_ws.cell(row_index, 2, sprint_wd); c.fill = INPUT_FILL; c.number_format = "0"
    cap_ws.cell(row_index, 3, f"FILTER 15-day sprint Tue → Mon. {SPRINT_NAME} = {sprint_start} → {sprint_end}.")
    row_index += 1
    p_rows["holidays"] = row_index
    cap_ws.cell(row_index, 1, "Company holidays in sprint")
    c = cap_ws.cell(row_index, 2, holiday_count); c.fill = INPUT_FILL; c.number_format = "0"
    cap_ws.cell(row_index, 3, f"India calendar 2026 — {holiday_descr}")
    row_index += 1
    p_rows["effwd"] = row_index
    cap_ws.cell(row_index, 1, "Effective working days (sprint − company holidays)")
    cap_ws.cell(row_index, 2, f"=B{p_rows['wd']}-B{p_rows['holidays']}").number_format = "0"
    cap_ws.cell(row_index, 3, "Used as the per-person calendar for availability")
    row_index += 1
    p_rows["hpd"] = row_index
    cap_ws.cell(row_index, 1, "Hours per working day (nominal)")
    c = cap_ws.cell(row_index, 2, 8); c.fill = INPUT_FILL; c.number_format = "0.00"
    cap_ws.cell(row_index, 3, "Standard working day")
    row_index += 1
    p_rows["focus"] = row_index
    cap_ws.cell(row_index, 1, "Focus factor (productive fraction)")
    c = cap_ws.cell(row_index, 2, 0.9); c.fill = INPUT_FILL; c.number_format = "0.00"
    cap_ws.cell(row_index, 3, "90% — what SLT expects (sprint-analysis default)")
    row_index += 1
    p_rows["misc"] = row_index
    cap_ws.cell(row_index, 1, "Misc buffer (meetings, scrums, code review, …)")
    c = cap_ws.cell(row_index, 2, MISC_BUFFER_PCT); c.fill = INPUT_FILL; c.number_format = "0.0%"
    cap_ws.cell(row_index, 3, f"{int(MISC_BUFFER_PCT*100)}% blanket reservation deducted from Available to get Net Available")
    row_index += 1
    p_rows["eff"] = row_index
    cap_ws.cell(row_index, 1, "Effective hours per FTE per sprint")
    cap_ws.cell(row_index, 2, f"=B{p_rows['effwd']}*B{p_rows['hpd']}*B{p_rows['focus']}").number_format = "0.00"
    cap_ws.cell(row_index, 3, "= effective days × hrs/day × focus factor")
    row_index += 1
    p_rows["fte"] = row_index
    cap_ws.cell(row_index, 1, "Allocated FTE on FILTER")
    c = cap_ws.cell(row_index, 2, len(CORE_TEAM_ORDER)); c.fill = INPUT_FILL; c.number_format = "0"
    cap_ws.cell(row_index, 3, "Count of names in the per-person availability table")
    row_index += 1
    p_rows["ideal"] = row_index
    cap_ws.cell(row_index, 1, "Ideal capacity for the sprint (hrs)")
    cap_ws.cell(row_index, 2, f"=B{p_rows['eff']}*B{p_rows['fte']}").number_format = "0.00"
    cap_ws.cell(row_index, 3, "= effective hrs/FTE × allocated FTE")

    row_index += 2
    cap_ws.cell(row_index, 1, f"Capacity vs Actual Delivery — {SPRINT_NAME}").font = Font(bold=True)
    cap_ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=7)
    row_index += 1
    for c, h in enumerate(["Metric", "Hours", "% of Ideal", "Rating", "Notes"], 1):
        cell = cap_ws.cell(row_index, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row_index += 1

    pp_total_row_placeholder = "{TR}"
    rating_formula = (
        '=IF(C{row}="","",IF(C{row}>=0.85,"On Target",'
        'IF(C{row}>=0.7,"Below Target","Significantly Below")))'
    )
    metrics = [
        ("Actual Availability", f"='Per-person Delivery'!F{pp_total_row_placeholder}", None,
         "Sum of per-person availability (Ideal − leaves)"),
        (f"Misc Buffer ({int(MISC_BUFFER_PCT*100)}%)",
         f"='Per-person Delivery'!H{pp_total_row_placeholder}", None,
         "Reserved for meetings, scrums, code reviews"),
        ("Net Available (Available − Misc Buffer)",
         f"='Per-person Delivery'!I{pp_total_row_placeholder}", None,
         "Realistic delivery capacity — used as the denominator for Completed %"),
        ("Completed", f"='Per-person Delivery'!B{pp_total_row_placeholder}",
         "=IFERROR(B{row}/B{net_row},\"\")",
         "Sum of in-sprint timeSpent (Completed bucket)"),
        ("Not Completed (Spillover)", f"='Per-person Delivery'!C{pp_total_row_placeholder}",
         "=IFERROR(B{row}/B{net_row},\"\")",
         "Original estimate for spillover (carried into next sprint)"),
        ("Removed", f"='Per-person Delivery'!D{pp_total_row_placeholder}",
         "=IFERROR(B{row}/B{net_row},\"\")",
         "Original estimate for issues removed mid-sprint"),
    ]
    metric_start_row = row_index
    net_row = metric_start_row + 2  # third metric is Net Available
    for label, hrs_formula, pct_formula, notes in metrics:
        cap_ws.cell(row_index, 1, label)
        cap_ws.cell(row_index, 2, hrs_formula).number_format = "0.00"
        if pct_formula is not None:
            cap_ws.cell(row_index, 3, pct_formula.format(row=row_index, net_row=net_row)).number_format = "0.0%"
        cap_ws.cell(row_index, 4, rating_formula.format(row=row_index))
        cap_ws.cell(row_index, 5, notes)
        row_index += 1

    row_index += 1
    cap_ws.cell(row_index, 1, "Completed by Priority").font = Font(bold=True)
    cap_ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=7)
    row_index += 1
    cap_ws.cell(
        row_index, 1,
        "Tickets (parents) = unique parent items (Story / Task / Defect / Escape Defect) — subtasks roll "
        "up to their parent so 1 parent + 3 subtasks = 1 ticket. "
        "Tickets (incl. subtasks) = raw record count incl. every Dev / QA / Automation / Review / "
        "Test Case Creation / Debug subtask. "
        "Hours = sum across parents and subtasks; matches the core-team Completed hours on the "
        "Per-person Delivery sheet.",
    ).font = Font(italic=True, color="FF555555")
    cap_ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=7)
    cap_ws.row_dimensions[row_index].height = 28
    cap_ws.cell(row_index, 1).alignment = Alignment(wrap_text=True, vertical="top")
    row_index += 1
    pri_headers = [
        "Priority",
        "Tickets (parents)",
        "Tickets (incl. subtasks)",
        "Hours (incl. subtasks)",
        "% of Completed Hrs",
        "% of Tickets (parents)",
        "Notes",
    ]
    for c, h in enumerate(pri_headers, 1):
        cell = cap_ws.cell(row_index, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row_index += 1
    pri_first_row = row_index
    total_completed_tickets = sum(v["count"] for v in completed_priority.values())
    total_completed_tickets_all = sum(v["count_all"] for v in completed_priority.values())
    total_completed_hours = sum(v["hours"] for v in completed_priority.values())
    for p in PRIORITY_ORDER:
        v = completed_priority[p]
        if v["count"] == 0 and v["count_all"] == 0 and v["hours"] == 0.0 and p != "P1":
            continue
        cap_ws.cell(row_index, 1, p)
        cap_ws.cell(row_index, 2, v["count"]).number_format = "0"
        cap_ws.cell(row_index, 3, v["count_all"]).number_format = "0"
        cap_ws.cell(row_index, 4, v["hours"]).number_format = "0.00"
        cap_ws.cell(
            row_index, 5,
            f'=IFERROR(D{row_index}/{total_completed_hours or 1},"")',
        ).number_format = "0.0%"
        cap_ws.cell(
            row_index, 6,
            f'=IFERROR(B{row_index}/{total_completed_tickets or 1},"")',
        ).number_format = "0.0%"
        if p == "P0":
            cap_ws.cell(row_index, 7, "Critical / production halting")
        elif p == "P1":
            cell = cap_ws.cell(row_index, 7, "P1 — high priority (see P1 Escalation tab for escalations)")
            cell.font = Font(bold=True, color="FFC00000")
        elif p == "P2":
            cap_ws.cell(row_index, 7, "Major")
        elif p == "P3":
            cap_ws.cell(row_index, 7, "Standard")
        elif p == "P4":
            cap_ws.cell(row_index, 7, "Minor / nice-to-have")
        else:
            cap_ws.cell(row_index, 7, "Priority not set in Jira")
        row_index += 1
    cap_ws.cell(row_index, 1, "TOTAL").font = HEADER_FONT
    cap_ws.cell(row_index, 2, total_completed_tickets).number_format = "0"
    cap_ws.cell(row_index, 3, total_completed_tickets_all).number_format = "0"
    cap_ws.cell(row_index, 4, total_completed_hours).number_format = "0.00"
    cap_ws.cell(row_index, 5, 1.0).number_format = "0.0%"
    cap_ws.cell(row_index, 6, 1.0).number_format = "0.0%"
    cap_ws.cell(row_index, 7, "Sum of Completed bucket (core roster)")
    row_index += 2

    cap_ws.cell(row_index, 1, "Sprint Completion (spillover = Not Completed)").font = Font(bold=True)
    cap_ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=7)
    row_index += 1
    sc_headers = ["Metric", "Tickets", "Planned Hrs (OE)", "% by Tickets", "% by Hrs", "Notes"]
    for c, h in enumerate(sc_headers, 1):
        cell = cap_ws.cell(row_index, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row_index += 1

    in_sprint_tickets = n_completed + n_spillover
    in_sprint_oe = round(completed_oe_hrs + spillover_oe_hrs, 2)
    planned_in_sprint_tickets = n_planned_completed + n_planned_spillover
    planned_in_sprint_oe = round(planned_completed_oe + planned_spillover_oe, 2)

    def safe_pct(num: float, den: float) -> str:
        if not den:
            return "—"
        return f"=IFERROR({num}/{den},\"\")"

    sc_rows = [
        ("Completed (closed in sprint)", n_completed, completed_oe_hrs,
         safe_pct(n_completed, in_sprint_tickets),
         safe_pct(completed_oe_hrs, in_sprint_oe),
         "In-sprint deliverables"),
        ("Spillover (not completed → carry-over)", n_spillover, spillover_oe_hrs,
         safe_pct(n_spillover, in_sprint_tickets),
         safe_pct(spillover_oe_hrs, in_sprint_oe),
         "Original estimate carried to next sprint"),
        ("Removed (mid-sprint scope cut)", n_removed, removed_oe_hrs,
         "—", "—",
         "Excluded from completion ratio (out of scope)"),
        ("Mid-sprint additions (core)", n_added_total_core, "",
         "—", "—",
         "Late-arriving work (Jira issueKeysAddedDuringSprint)"),
    ]
    for label, tk, oe, pct_t, pct_h, note in sc_rows:
        cap_ws.cell(row_index, 1, label)
        cap_ws.cell(row_index, 2, tk).number_format = "0"
        cap_ws.cell(row_index, 3, oe if oe != "" else None)
        if isinstance(oe, (int, float)):
            cap_ws.cell(row_index, 3).number_format = "0.00"
        cap_ws.cell(row_index, 4, pct_t if pct_t != "—" else "—")
        if pct_t != "—":
            cap_ws.cell(row_index, 4).number_format = "0.0%"
        cap_ws.cell(row_index, 5, pct_h if pct_h != "—" else "—")
        if pct_h != "—":
            cap_ws.cell(row_index, 5).number_format = "0.0%"
        cap_ws.cell(row_index, 6, note)
        row_index += 1
    row_index += 1

    sprint_completion_count = (n_completed / in_sprint_tickets) if in_sprint_tickets else 0
    sprint_completion_hrs = (completed_oe_hrs / in_sprint_oe) if in_sprint_oe else 0
    plan_completion_count = (n_planned_completed / planned_in_sprint_tickets) if planned_in_sprint_tickets else 0
    plan_completion_hrs = (planned_completed_oe / planned_in_sprint_oe) if planned_in_sprint_oe else 0

    def rating_label(value: float) -> tuple[str, PatternFill, Font]:
        if value >= 0.85:
            return "On Target", GREEN, Font(color="FF006100", bold=True)
        if value >= 0.7:
            return "Below Target", AMBER, Font(color="FF9C5700", bold=True)
        return "Significantly Below", RED, Font(color="FF9C0006", bold=True)

    summary_rows = [
        ("Sprint completion rate (tickets)", sprint_completion_count,
         "Completed ÷ (Completed + Spillover)"),
        ("Sprint completion rate (planned hrs)", sprint_completion_hrs,
         "Σ OE Completed ÷ Σ OE (Completed + Spillover)"),
        ("Planned completion rate (tickets, excludes mid-sprint adds)", plan_completion_count,
         "Originally-committed tickets only"),
        ("Planned completion rate (planned hrs, excludes mid-sprint adds)", plan_completion_hrs,
         "Originally-committed planned hours only"),
    ]
    for label, value, note in summary_rows:
        cap_ws.cell(row_index, 1, label).font = Font(bold=True)
        v_cell = cap_ws.cell(row_index, 2, value)
        v_cell.number_format = "0.0%"
        rating, fill, font = rating_label(value)
        r_cell = cap_ws.cell(row_index, 3, rating)
        r_cell.fill = fill
        r_cell.font = font
        cap_ws.cell(row_index, 4, note).alignment = Alignment(wrap_text=True)
        cap_ws.merge_cells(start_row=row_index, start_column=4, end_row=row_index, end_column=7)
        row_index += 1
    row_index += 1

    p1_total = len(p1_escalation_recs)
    p1_open = sum(1 for r in p1_escalation_recs if (r.get("status_category") or "") != "done")
    cap_ws.cell(row_index, 1, "P1 Escalation snapshot").font = P1_ESC_TITLE_FONT
    cap_ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=7)
    row_index += 1
    p1_summary = (
        f"{p1_total} ticket(s) in {SPRINT_NAME} match 'P1 Escalation' in the title — "
        f"{p1_total - p1_open} closed, {p1_open} still open. See the 'P1 Escalation' tab."
    )
    cell = cap_ws.cell(row_index, 1, p1_summary)
    cell.font = Font(bold=True, color="FFC00000")
    cap_ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=7)
    row_index += 2

    completed_added = sum(1 for r in completed_recs if r.get("added_during_sprint"))
    not_completed_added = sum(1 for r in not_completed_recs if r.get("added_during_sprint"))
    removed_added = sum(1 for r in removed_recs if r.get("added_during_sprint"))
    cap_ws.cell(row_index, 1, "Mid-sprint additions snapshot").font = Font(bold=True, color="FF7F6000", size=12)
    cap_ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=7)
    row_index += 1
    add_summary = (
        f"{n_added_total_core} ticket(s) were added to {SPRINT_NAME} after it started "
        f"(issueKeysAddedDuringSprint) — Completed {completed_added}, "
        f"Spillover {not_completed_added}, Removed {removed_added}. "
        f"Highlighted in amber on every detail sheet (Added Mid-Sprint = YES)."
    )
    cell = cap_ws.cell(row_index, 1, add_summary)
    cell.font = Font(bold=True, color="FF7F6000")
    cap_ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=7)
    cap_ws.row_dimensions[row_index].height = 36
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    row_index += 2

    cap_ws.cell(row_index, 1, "Assumptions & caveats").font = Font(bold=True)
    row_index += 1
    notes_lines = [
        f"• Sprint window from Jira ({sprint_meta.get('api_self', '/rest/agile/1.0/sprint/'+str(sprint_meta['sprint_id']))}): "
        f"{sprint_start} → {sprint_end} ({sprint_wd} Mon–Fri working days). Refresh with "
        "fetch_filter_sprint_meta.py + fetch_filter_sprint_issues.py.",
        "• Available (hrs) = (effective working days − leave days) × 8 × focus factor, where "
        "effective working days = sprint working days − company holidays falling inside the sprint. "
        "Per-person Delivery shows leave days in Column G to explain the delta from a full sprint.",
        f"• Net Available = Available − {int(MISC_BUFFER_PCT*100)}% Misc Buffer for meetings, scrums, "
        "code review, etc. Completed/Not Completed/Removed % of capacity use Net Available as the denominator.",
        "• Time Spent on Completed sheet uses the in-sprint hybrid worklog rule "
        "(sprint-analysis.md): single-sprint = timespent; multi-sprint = SUM of worklogs whose "
        "started ∈ [sprintStart, sprintEnd].",
        "• Not Completed and Removed totals use Original Estimate.",
        "• Removed list comes from the Sprint Report puntedIssues (greenhopper API).",
        "• Core roster only: assignees outside Securly-PjM-Skills.md Part 2 are dropped from per-person + detail sheets.",
        "• P1 Escalation tab is filtered to Issue Type = Escape Defect AND Priority = P1 AND title contains 'P1 Escalation'.",
        "• Completed by Priority — Tickets (parents) rolls subtasks up to their parent (each work item "
        "counted once at its parent priority); Tickets (incl. subtasks) shows the raw record count for "
        "comparison. The Hours column sums every record (parent + subtasks) and matches the core-team "
        "Completed total.",
        "• Spillover = Not Completed bucket (tickets carried into the next sprint).",
        "• Sprint completion rate uses Completed ÷ (Completed + Spillover); Removed tickets are excluded "
        "(treated as scope cut). 'Planned' rate also excludes mid-sprint additions.",
        "• All detail sheets (Completed, Not Completed, Removed) highlight tickets added after sprint start "
        "(Jira sprint-report issueKeysAddedDuringSprint) in amber, with an 'Added Mid-Sprint = YES' tag column.",
        "• Ratings: ≥85% On Target (green), 70–85% Below Target (amber), <70% Significantly Below (red).",
    ]
    for line in notes_lines:
        cell = cap_ws.cell(row_index, 1, line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cap_ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=7)
        cap_ws.row_dimensions[row_index].height = 32
        row_index += 1

    cap_ws.column_dimensions["A"].width = 44
    cap_ws.column_dimensions["B"].width = 14
    cap_ws.column_dimensions["C"].width = 32
    cap_ws.column_dimensions["D"].width = 18
    cap_ws.column_dimensions["E"].width = 14
    cap_ws.column_dimensions["F"].width = 14
    cap_ws.column_dimensions["G"].width = 36

    pp_ws = wb.create_sheet("Per-person Delivery")
    pp_headers = [
        "Assignee",                                              # A
        "Completed (hrs)",                                       # B
        "Not Completed (hrs)",                                   # C
        "Removed (hrs)",                                         # D
        "Efficiency",                                            # E (vs Net Available)
        "Available (hrs)",                                       # F
        "Leave Days (sprint)",                                   # G — explains Available delta
        f"Misc Buffer {int(MISC_BUFFER_PCT*100)}% (hrs)",        # H — meetings / scrums
        "Net Available (hrs)",                                   # I = F − H
    ]
    for c, h in enumerate(pp_headers, 1):
        cell = pp_ws.cell(1, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for i, name in enumerate(CORE_TEAM_ORDER, 2):
        pp_ws.cell(i, 1, name)
        pp_ws.cell(i, 2, completed_hrs_by[name]).number_format = "0.00"
        pp_ws.cell(i, 3, notcomp_hrs_by[name]).number_format = "0.00"
        pp_ws.cell(i, 4, removed_hrs_by[name]).number_format = "0.00"
        pp_ws.cell(i, 5, f'=IFERROR(B{i}/I{i},"")').number_format = "0.0%"
        pp_ws.cell(i, 6, avail_hours[name]).number_format = "0.00"
        pp_ws.cell(i, 7, int(leave_days.get(name, 0))).number_format = "0"
        pp_ws.cell(i, 8, f"=F{i}*{MISC_BUFFER_PCT}").number_format = "0.00"
        pp_ws.cell(i, 9, f"=F{i}-H{i}").number_format = "0.00"
    last = 1 + len(CORE_TEAM_ORDER)
    tr = last + 1
    pp_ws.cell(tr, 1, "TOTAL").font = HEADER_FONT
    pp_ws.cell(tr, 2, f"=SUM(B2:B{last})").number_format = "0.00"
    pp_ws.cell(tr, 3, f"=SUM(C2:C{last})").number_format = "0.00"
    pp_ws.cell(tr, 4, f"=SUM(D2:D{last})").number_format = "0.00"
    pp_ws.cell(tr, 5, f'=IFERROR(B{tr}/I{tr},"")').number_format = "0.0%"
    pp_ws.cell(tr, 6, f"=SUM(F2:F{last})").number_format = "0.00"
    pp_ws.cell(tr, 7, f"=SUM(G2:G{last})").number_format = "0"
    pp_ws.cell(tr, 8, f"=SUM(H2:H{last})").number_format = "0.00"
    pp_ws.cell(tr, 9, f"=SUM(I2:I{last})").number_format = "0.00"
    for col_letter, w in zip("ABCDEFGHI", [22, 14, 16, 14, 12, 14, 14, 18, 16]):
        pp_ws.column_dimensions[col_letter].width = w
    pp_ws.row_dimensions[1].height = 32
    pp_ws.freeze_panes = "B2"

    for placeholder_row in range(metric_start_row, metric_start_row + len(metrics)):
        cell = cap_ws.cell(placeholder_row, 2)
        if isinstance(cell.value, str):
            cell.value = cell.value.replace(pp_total_row_placeholder, str(tr))

    n_added_completed = write_detail_sheet(
        wb.create_sheet("Completed"),
        completed_recs,
        ("assignee", "key"),
        flag_mid_sprint=True,
    )
    n_added_not_completed = write_detail_sheet(
        wb.create_sheet("Not Completed"),
        not_completed_recs,
        ("assignee", "key"),
        flag_mid_sprint=True,
    )
    n_added_removed = write_detail_sheet(
        wb.create_sheet("Removed"),
        removed_recs,
        ("assignee", "key"),
        flag_mid_sprint=True,
    )
    p1_ws = wb.create_sheet("P1 Escalation")
    p1_ws.sheet_properties.tabColor = "C00000"
    write_p1_escalation_sheet(p1_ws, p1_escalation_recs, sprint_meta)

    wb._sheets = [
        wb["Capacity vs Delivery"],
        wb["Per-person Delivery"],
        wb["Completed"],
        wb["Not Completed"],
        wb["Removed"],
        wb["P1 Escalation"],
    ]

    out = ROOT / "FILTER Sprint Hanna Analysis.xlsx"
    try:
        wb.save(out)
    except PermissionError:
        from datetime import datetime as _dt
        suffix = _dt.now().strftime("%H%M%S")
        out = ROOT / f"FILTER Sprint Hanna Analysis.{suffix}.xlsx"
        print(f"Primary file is open in Excel — saving to {out.name} instead.")
        wb.save(out)
    print(f"Wrote {out}")
    print(
        f"  completed_rows={len(completed_recs)} "
        f"not_completed_rows={len(not_completed_recs)} "
        f"removed_rows={len(removed_recs)}"
    )
    print(f"  total Completed (in-sprint hrs) on Per-person: {sum(completed_hrs_by.values()):.2f}")
    print(f"  total Not Completed (orig est hrs):           {sum(notcomp_hrs_by.values()):.2f}")
    print(f"  total Removed (orig est hrs):                 {sum(removed_hrs_by.values()):.2f}")
    print(f"  Actual Availability (sum of per-person):      {sum(avail_hours.values()):.2f}")
    print(
        f"  Company holidays in sprint: {holiday_count} "
        f"(effective working days = {eff_sprint_wd} of {sprint_wd}). {holiday_descr}"
    )
    print("  Completed by priority (parents / incl. subtasks / hrs):")
    for p in PRIORITY_ORDER:
        v = completed_priority[p]
        if v["count"] or v["count_all"] or v["hours"]:
            print(
                f"    {p:<7} parents={v['count']:>4}  incl.subtasks={v['count_all']:>4}  "
                f"hrs={v['hours']:>8.2f}"
            )
    print(
        f"  Sprint completion (tickets):   "
        f"{n_completed}/{in_sprint_tickets} "
        f"= {sprint_completion_count*100:.1f}%   "
        f"(planned: {n_planned_completed}/{planned_in_sprint_tickets} "
        f"= {plan_completion_count*100:.1f}%)"
    )
    print(
        f"  Sprint completion (planned hrs): "
        f"{completed_oe_hrs:.2f} / {in_sprint_oe:.2f} "
        f"= {sprint_completion_hrs*100:.1f}%   "
        f"(planned-only: {planned_completed_oe:.2f} / {planned_in_sprint_oe:.2f} "
        f"= {plan_completion_hrs*100:.1f}%)"
    )
    print(
        f"  Mid-sprint additions (core): {n_added_total_core} "
        f"(Completed {completed_added}, Spillover {not_completed_added}, Removed {removed_added})"
    )
    print(f"  P1 Escalation tickets: total={len(p1_escalation_recs)}")


if __name__ == "__main__":
    main()
